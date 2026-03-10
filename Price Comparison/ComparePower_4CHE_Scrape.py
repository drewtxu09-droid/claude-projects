"""
ComparePower_4CHE_Scrape.py
Scrapes electricity rates from comparepower.com for zip code 76051,
exports to Excel sorted by price, and reports 4Change Energy's ranking.
"""

import time
import re
import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

ZIP_CODE = "76051"
ONEDRIVE_ALERT_FOLDER = r"C:\Users\XV1S\OneDrive - Vistra Corp\ComparePower Alerts"


def write_alert_file(rank, plan, price, bill, leader, leader_price, leader_bill):
    """Write a JSON trigger file to OneDrive so Power Automate sends a Teams alert."""
    os.makedirs(ONEDRIVE_ALERT_FOLDER, exist_ok=True)
    filename = f"alert_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(ONEDRIVE_ALERT_FOLDER, filename)
    payload = {
        "rank": rank,
        "plan": plan,
        "price": round(price, 2),
        "bill": round(bill, 2),
        "leader": leader,
        "leader_price": round(leader_price, 2),
        "leader_bill": round(leader_bill, 2),
        "timestamp": datetime.now().strftime("%Y-%m-%d %I:%M %p"),
    }
    with open(filepath, "w") as f:
        json.dump(payload, f, indent=2)
    print(f"  Alert file written to OneDrive: {filename}")

def make_output_filename():
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    hour = now.strftime("%I").lstrip("0") or "12"
    minute = now.strftime("%M")
    ampm = now.strftime("%p").lower()
    return f"4CHE/ComparePower_4CHE_{ZIP_CODE}_{date_str}_{hour}{minute}{ampm}.xlsx"


def setup_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def click_all_plans(driver):
    """Click the 'All Plans' button/link if it appears after zip submission."""
    for _ in range(8):
        try:
            btn = driver.find_element(By.XPATH, "//*[normalize-space(text())='All Plans' or normalize-space(text())='ALL PLANS']")
            if btn.is_displayed():
                btn.click()
                print("  Clicked 'All Plans'.")
                time.sleep(2)
                return True
        except Exception:
            pass
        time.sleep(1)
    print("  'All Plans' button not found — proceeding with whatever loaded.")
    return False


def enter_zip_and_load(driver):
    print(f"Opening comparepower.com for zip {ZIP_CODE}...")
    driver.get("https://www.comparepower.com/")
    wait = WebDriverWait(driver, 20)

    # The page has two identical zip inputs: index 0 is hidden (desktop hero),
    # index 1 is the visible one. Use the visible input.
    print("  Waiting for zip input...")
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input.vc-zipcode")))
    all_inputs = driver.find_elements(By.CSS_SELECTOR, "input.vc-zipcode")
    zip_input = next((x for x in all_inputs if x.is_displayed()), all_inputs[-1])
    zip_input.click()
    zip_input.clear()
    zip_input.send_keys(ZIP_CODE)
    time.sleep(0.5)

    # Click the visible "Click to Compare" button
    print("  Clicking 'Click to Compare'...")
    all_btns = driver.find_elements(By.CSS_SELECTOR, "button.cbutton")
    compare_btn = next((b for b in all_btns if b.is_displayed()), all_btns[-1])
    compare_btn.click()
    time.sleep(3)

    # Handle the "All Plans" prompt that appears after zip submission
    click_all_plans(driver)

    # Wait for plan cards to appear (orders.comparepower.com uses Quasar q-card)
    print("  Waiting for plans to load...")
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.q-card")))
        print("  Plans detected.")
    except Exception:
        print("  Warning: Could not confirm plans loaded. Proceeding anyway.")

    # Scroll to trigger lazy-loaded content
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(8):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(2)


# Known slug → display name overrides for short/ambiguous slugs
PROVIDER_SLUG_MAP = {
    "apge": "AP Gas & Electric",
    "txu": "TXU Energy",
    "txu_energy": "TXU Energy",
    "constellation": "Constellation",
    "nrg": "NRG Energy",
}


def slug_to_provider(slug):
    """Convert a provider image filename slug to a display name."""
    name = re.sub(r"\.(svg|png|jpg|webp)$", "", slug, flags=re.I)
    if name in PROVIDER_SLUG_MAP:
        return PROVIDER_SLUG_MAP[name]
    # Replace separators, title-case each word
    name = name.replace("_", " ").replace("-", " ")
    parts = []
    for word in name.split():
        if word and word[0].isdigit():
            parts.append(word[0] + word[1:].capitalize())
        else:
            parts.append(word.capitalize())
    return " ".join(parts)


def parse_price(text):
    """Extract a float price in cents/kWh from a string."""
    if not text:
        return None
    raw = re.sub(r"[^\d.]", "", text)
    if raw:
        try:
            val = float(raw)
            if val < 1:
                val = round(val * 100, 4)
            return val
        except ValueError:
            pass
    return None


def parse_card(card):
    """Parse a single q-card Selenium element into a plan dict."""
    text = card.text.strip()
    if not text or "effective rate per kWh" not in text:
        return None

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # --- Provider from logo image ---
    provider = ""
    imgs = card.find_elements(By.TAG_NAME, "img")
    for img in imgs:
        src = img.get_attribute("src") or ""
        if "comparepower.com/images" in src or "assets.comparepower" in src:
            filename = src.split("/")[-1]
            provider = slug_to_provider(filename)
            break

    # Fallback: extract from "Provider | PUCT # XXXXX" line (Type A)
    if not provider:
        for line in reversed(lines):
            if "PUCT" in line.upper() and "|" in line:
                provider = re.sub(r"\s*\|?\s*PUCT\s*#?\s*\d+.*", "", line, flags=re.I).strip()
                if provider:
                    break

    if not provider:
        return None

    # --- Price: line immediately before "effective rate per kWh" ---
    price = None
    price_idx = -1
    for i, line in enumerate(lines):
        if "effective rate per kwh" in line.lower() and i > 0:
            raw = re.sub(r"[^\d.]", "", lines[i - 1])
            try:
                price = float(raw)
                if price < 1:
                    price = round(price * 100, 4)
                price_idx = i - 1
            except ValueError:
                pass
            break

    if price is None or not (1 < price < 100):
        return None

    # --- Plan name ---
    plan_name = ""
    monthly_bill_idx = next(
        (i for i, l in enumerate(lines) if "monthly bill estimate" in l.lower()), -1
    )
    if monthly_bill_idx != -1 and monthly_bill_idx + 1 < len(lines):
        plan_name = lines[monthly_bill_idx + 1]
    else:
        for i in range(price_idx - 1, max(price_idx - 5, -1), -1):
            candidate = lines[i]
            if re.search(r"featured|popular|100%|renewable|new!?$|sale!?$", candidate, re.I):
                continue
            if len(candidate) > 3:
                plan_name = candidate
                break

    # --- Contract ---
    contract = ""
    for line in lines:
        m = re.search(r"(\d+)[\s-]month", line, re.I)
        if m:
            contract = f"{m.group(1)} months"
            break
        if "month to month" in line.lower() or "no contract" in line.lower():
            contract = "Month-to-month"
            break

    # --- Monthly bill estimate ---
    # Type B: line immediately before "monthly bill estimate for 1000 kWh usage"
    # Type A: embedded in "($XX.XX – 1000 kWh)" line after the price
    bill_est = None
    if monthly_bill_idx != -1 and monthly_bill_idx > 0:
        raw = re.sub(r"[^\d.]", "", lines[monthly_bill_idx - 1])
        try:
            bill_est = float(raw)
        except ValueError:
            pass
    if bill_est is None:
        # Type A fallback: find "($XX.XX" pattern
        for line in lines:
            m = re.search(r"\(\$?([\d.]+)", line)
            if m:
                try:
                    bill_est = float(m.group(1))
                    break
                except ValueError:
                    pass

    return {
        "Provider": provider,
        "Plan Name": plan_name,
        "Price (¢/kWh)": price,
        "Bill Est. ($)": bill_est,
        "Contract": contract,
    }


def scrape_rates(driver):
    print("Scraping rate data...")
    plans = []

    cards = driver.find_elements(By.CSS_SELECTOR, "div.q-card")
    if cards:
        print(f"  Found {len(cards)} q-card elements. Parsing...")
        seen = set()
        for card in cards:
            try:
                plan = parse_card(card)
                if plan:
                    key = (plan["Provider"], plan["Plan Name"], plan["Price (¢/kWh)"])
                    if key not in seen:
                        seen.add(key)
                        plans.append(plan)
            except Exception:
                continue

        if plans:
            print(f"  Extracted {len(plans)} unique plans from cards.")
            return plans

    print("  No q-cards with rate data found. Attempting page source parse...")
    return parse_from_page_source(driver)


def parse_from_page_source(driver):
    """Fallback: extract structured data from page source using regex."""
    source = driver.page_source
    plans = []

    json_matches = re.findall(r'"provider"\s*:\s*"([^"]+)".*?"price"\s*:\s*"?(\d+\.?\d*)"?', source, re.DOTALL)
    for provider, price in json_matches:
        plans.append({
            "Provider": provider,
            "Plan Name": "",
            "Price (¢/kWh)": float(price),
            "Contract": "",
        })

    return plans


def build_excel(plans, output_file):
    if not plans:
        print("\nNo rate data was captured. The site may require interaction or has changed its structure.")
        return

    df = pd.DataFrame(plans)
    df = df.drop_duplicates(subset=["Provider", "Plan Name", "Price (¢/kWh)"])
    df = df.sort_values("Bill Est. ($)", na_position="last").reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df) + 1))

    # Comparison columns vs #1
    top_price = df.iloc[0]["Price (¢/kWh)"]
    top_bill  = df.iloc[0]["Bill Est. ($)"]
    df["vs #1 Price (¢/kWh)"] = (df["Price (¢/kWh)"] - top_price).round(2)
    df["vs #1 Bill ($)"]      = (df["Bill Est. ($)"]  - top_bill).round(2)

    target_mask = df["Provider"].str.contains("4Change", case=False, na=False)
    target_rows = df[target_mask]

    print(f"\n{'='*60}")
    print(f"  Total plans found: {len(df)}")
    print(f"{'='*60}")
    print(f"\n  #1 Cheapest: {df.iloc[0]['Provider']} — {df.iloc[0]['Price (¢/kWh)']:.2f}¢/kWh  (${top_bill:.2f}/mo)")

    if target_rows.empty:
        print("\n  4Change Energy: NOT FOUND in results.")
    else:
        target_rank  = int(target_rows.iloc[0]["Rank"])
        target_price = float(target_rows.iloc[0]["Price (¢/kWh)"])
        target_bill  = float(target_rows.iloc[0]["Bill Est. ($)"])
        target_plan  = target_rows.iloc[0]["Plan Name"]
        print(f"\n  4Change Energy rank: #{target_rank} ({target_plan or 'N/A'}) — {target_price:.2f}¢/kWh  (${target_bill:.2f}/mo)")

        if target_rank == 1:
            print("  4Change Energy IS the cheapest provider! No alert sent.")
        else:
            diff_price = target_price - top_price
            diff_bill  = target_bill  - top_bill
            print(f"  4Change is NOT #1. Cheapest is {df.iloc[0]['Provider']} at {top_price:.2f}¢/kWh.")
            print(f"  4Change pays +{diff_price:.2f}¢/kWh more  (+${diff_bill:.2f}/mo more than #1).")
            write_alert_file(
                rank=target_rank,
                plan=target_plan or "N/A",
                price=target_price,
                bill=target_bill,
                leader=df.iloc[0]["Provider"],
                leader_price=top_price,
                leader_bill=top_bill,
            )

    print(f"{'='*60}\n")

    df.to_excel(output_file, index=False, sheet_name="Rates")

    wb = load_workbook(output_file)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    target_fill = PatternFill("solid", fgColor="E8D5F5")  # light purple
    top_fill = PatternFill("solid", fgColor="C6EFCE")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        rank_val = row[0].value
        provider_val = str(row[1].value) if row[1].value else ""
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if rank_val == 1:
                cell.fill = top_fill
            elif "4change" in provider_val.lower():
                cell.fill = target_fill
        # Col G (index 6): vs #1 Price — format as ¢ with sign
        # Col H (index 7): vs #1 Bill — format as $ with sign
        if len(row) > 6:
            row[6].number_format = '+0.00"¢";-0.00"¢";"0¢"'
        if len(row) > 7:
            row[7].number_format = '+$#,##0.00;-$#,##0.00;"$0.00"'

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    wb.save(output_file)
    print(f"  Excel saved to: {output_file}")


def main():
    output_file = make_output_filename()
    driver = setup_driver()
    try:
        enter_zip_and_load(driver)
        plans = scrape_rates(driver)
        build_excel(plans, output_file)
        print("\nDone. Closing browser.")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
